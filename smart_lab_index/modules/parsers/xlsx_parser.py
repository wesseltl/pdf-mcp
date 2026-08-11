"""Excel XLSX parser module."""

from __future__ import annotations

from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from smart_lab_index.core.domain import DocumentContent, DocumentSource, TableContent
from smart_lab_index.core.modules import (
    FileAccess,
    ModuleCapability,
    ModuleManifest,
    ModuleType,
    NetworkAccess,
    ParserModule,
)
from smart_lab_index.modules.parsers.common import (
    DocumentParseError,
    assess_table,
    clean_cell,
    preflight_office_archive,
    provenance,
    require_table_budget,
    source_extension,
)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class XlsxParser(ParserModule):
    manifest = ModuleManifest(
        module_id="parser.xlsx",
        name="XLSX Parser",
        version="0.2.0",
        module_type=ModuleType.PARSER,
        description="Extracts worksheet rows and cell provenance from Excel streams.",
        capabilities=(ModuleCapability("parser.document", "1.0.0"),),
        network_access=NetworkAccess.NONE,
        file_access=FileAccess.NONE,
    )

    def supports(self, source: DocumentSource) -> bool:
        return source_extension(source) == ".xlsx" or source.content_type == XLSX_CONTENT_TYPE

    def parse(self, source: DocumentSource, content: BinaryIO) -> DocumentContent:
        workbook = None
        try:
            preflight_office_archive(content, "XLSX document")
            workbook = load_workbook(
                content,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            tables = []
            table_rows = 0
            table_cells = 0
            for sheet_index, sheet in enumerate(workbook.worksheets):
                rows = []
                cell_references = []
                for worksheet_row, row in enumerate(sheet.iter_rows(), start=1):
                    values = tuple(clean_cell(cell.value) for cell in row)
                    if not any(values):
                        continue
                    table_rows += 1
                    table_cells += len(values)
                    require_table_budget(table_rows, table_cells, "XLSX document")
                    rows.append(values)
                    cell_references.append(tuple(
                        provenance(source, {
                            "sheet": sheet.title,
                            "row": worksheet_row,
                            "column": column_index,
                            "cell": f"{get_column_letter(column_index)}{worksheet_row}",
                        })
                        for column_index, _cell in enumerate(row, start=1)
                    ))
                if rows:
                    tables.append(TableContent(
                        index=sheet_index,
                        name=sheet.title,
                        rows=tuple(rows),
                        cell_provenance=tuple(cell_references),
                        metadata={
                            "sheet": sheet.title,
                            **assess_table(rows),
                        },
                    ))
            sheet_names = list(workbook.sheetnames)
        except DocumentParseError:
            raise
        except Exception as exc:
            raise DocumentParseError(f"XLSX parsing failed: {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()
        return DocumentContent(
            source_external_id=source.external_id,
            content_type=source.content_type,
            parser_module_id=self.manifest.module_id,
            parser_version=self.manifest.version,
            tables=tuple(tables),
            metadata={"sheet_names": sheet_names},
            warnings=() if tables else ("no non-empty Excel rows detected",),
        )
