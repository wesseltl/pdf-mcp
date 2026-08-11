"""Export extracted document tables to files people can use immediately."""
from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pdf_mcp import docx_extractor, extractor
from pdf_mcp.output_safety import spreadsheet_safe


def _source_type(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return "pdf"
    if ext == ".docx":
        return "docx"
    raise ValueError(f"unsupported input file type: {ext}")


def _output_type(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext in {".xlsx", ".csv", ".json"}:
        return ext[1:]
    raise ValueError(f"unsupported output file type: {ext}")


def extract_document_tables(
    path: str,
    merge_multipage: bool = True,
    source_name: str | None = None,
) -> dict:
    """Extract document tables without writing an output file.

    ``source_name`` changes only the human-readable source label included in exports. The input is
    still read exclusively from ``path``. This is useful for temporary-upload interfaces that
    should show the original filename rather than an internal temporary path.
    """
    kind = _source_type(path)
    if kind == "pdf":
        result = extractor.extract_tables(path, merge_multipage=merge_multipage)
    else:
        result = docx_extractor.extract_docx_tables(path)
    return {
        "source": source_name if source_name is not None else os.path.realpath(path),
        "source_type": kind,
        **result,
    }


def _extract_tables(path: str, merge_multipage: bool) -> dict:
    """Backward-compatible internal alias for older callers."""
    return extract_document_tables(path, merge_multipage=merge_multipage)


def _location(table: dict) -> str:
    if table.get("merged_from_pages"):
        pages = table["merged_from_pages"]
        return f"pages {pages[0]}-{pages[-1]}" if len(pages) > 1 else f"page {pages[0]}"
    if table.get("page") is not None:
        return f"page {table['page']}"
    if table.get("index") is not None:
        return f"table {table['index'] + 1}"
    return ""


def _warning_text(table: dict) -> str:
    return "; ".join(table.get("warnings") or [])


def _review_rows(tables: list[dict]) -> list[list]:
    return [
        ["table", "location", "rows", "columns", "looks_clean", "has_merged_cells", "warnings"],
        *[
            [
                i + 1,
                _location(table),
                table.get("n_rows", len(table.get("rows", []))),
                table.get("column_count"),
                table.get("looks_clean"),
                table.get("has_merged_cells", False),
                _warning_text(table),
            ]
            for i, table in enumerate(tables)
        ],
    ]


def _safe_sheet_title(title: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", title).strip() or "Table"
    cleaned = cleaned[:31]
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f" {n}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


def _write_xlsx(result: dict, output_path: str) -> None:
    wb = Workbook()
    review = wb.active
    review.title = "Review"
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)

    for row in _review_rows(result["tables"]):
        review.append([spreadsheet_safe(value) for value in row])
    for warning in result.get("warnings", []):
        review.append(["document warning", warning])
    for cell in review[1]:
        cell.font = header_font
    for row in review.iter_rows(min_row=2):
        if row[4].value is False:
            for cell in row:
                cell.fill = warning_fill
    review.freeze_panes = "A2"

    used = {"Review"}
    for i, table in enumerate(result["tables"]):
        sheet = wb.create_sheet(_safe_sheet_title(f"Table {i + 1}", used))
        sheet.append(["source", result["source"]])
        sheet.append(["location", _location(table)])
        sheet.append(["looks_clean", table.get("looks_clean")])
        sheet.append(["warnings", _warning_text(table)])
        sheet.append([])
        for row in table.get("rows", []):
            sheet.append([spreadsheet_safe(value) for value in row])
        for cell in sheet[1]:
            cell.font = header_font
    wb.save(output_path)


def _write_csv(result: dict, output_path: str) -> None:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["source", result["source"]])
        writer.writerow(["source_type", result["source_type"]])
        for warning in result.get("warnings", []):
            writer.writerow(["document_warning", warning])
        writer.writerow([])
        for i, table in enumerate(result["tables"]):
            writer.writerow([f"Table {i + 1}", _location(table)])
            writer.writerow(["looks_clean", table.get("looks_clean")])
            writer.writerow(["warnings", _warning_text(table)])
            writer.writerow([])
            writer.writerows(
                [spreadsheet_safe(value) for value in row]
                for row in table.get("rows", [])
            )
            writer.writerow([])


def _write_json(result: dict, output_path: str) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)
        f.write("\n")


def write_document_tables(result: dict, output_path: str) -> dict:
    """Write a previously extracted result to .xlsx, .csv, or .json."""
    resolved_output = extractor._check_path(output_path)
    os.makedirs(os.path.dirname(resolved_output) or ".", exist_ok=True)
    kind = _output_type(resolved_output)
    if kind == "xlsx":
        _write_xlsx(result, resolved_output)
    elif kind == "csv":
        _write_csv(result, resolved_output)
    else:
        _write_json(result, resolved_output)
    return {
        "input": result.get("source"),
        "output": resolved_output,
        "output_type": kind,
        "source_type": result.get("source_type"),
        "n_tables": result.get("n_tables", len(result.get("tables", []))),
        "tables_needing_review": sum(1 for t in result["tables"] if not t.get("looks_clean")),
        "warnings": result.get("warnings", []),
    }


def export_document_tables(input_path: str, output_path: str, merge_multipage: bool = True) -> dict:
    """Extract tables from a PDF or .docx and write them to .xlsx, .csv, or .json."""
    result = extract_document_tables(input_path, merge_multipage=merge_multipage)
    return write_document_tables(result, output_path)
