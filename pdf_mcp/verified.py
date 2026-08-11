"""Profile-driven extraction with deterministic validation and source evidence."""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pdf_mcp import __version__, docx_extractor, extractor
from pdf_mcp.output_safety import spreadsheet_safe
from pdf_mcp.profiles import load_profile, normalize_header, profile_sha256


RESULT_SCHEMA_VERSION = "1.0"


def _issue(severity: str, code: str, message: str, **context) -> dict:
    return {
        "severity": severity,
        "code": code,
        "message": message,
        **{key: value for key, value in context.items() if value is not None},
    }


def _package_version(package: str) -> str:
    try:
        return version(package)
    except PackageNotFoundError:  # pragma: no cover - editable installs include metadata
        return "unknown"


def _document(path: str) -> tuple[str, dict]:
    resolved = extractor._check_path(path)
    suffix = Path(resolved).suffix.lower()
    if suffix not in {".pdf", ".docx"}:
        raise ValueError("profile extraction supports only .pdf and .docx files")
    if not os.path.isfile(resolved):
        raise ValueError(f"document not found: {path}")
    digest = hashlib.sha256()
    size = 0
    with open(resolved, "rb") as document:
        for chunk in iter(lambda: document.read(1024 * 1024), b""):
            digest.update(chunk)
            size += len(chunk)
    if size == 0:
        raise ValueError("document is empty")
    source_type = suffix[1:]
    info = {
        "type": source_type,
        "bytes": size,
        "sha256": digest.hexdigest(),
    }
    if source_type == "pdf":
        info["pages"] = extractor.page_count(resolved)
    return resolved, info


def _source(table: dict) -> dict:
    source = {"table_index": table.get("index", 0)}
    if table.get("page") is not None:
        source["page"] = table["page"]
    if table.get("bbox") is not None:
        source["bbox"] = table["bbox"]
    return source


def _alias_lookup(columns: list[dict]) -> dict[str, str]:
    return {
        normalize_header(alias): column["name"]
        for column in columns
        for alias in column["aliases"]
    }


def _header_match(table: dict, profile_table: dict) -> dict:
    columns = profile_table["columns"]
    lookup = _alias_lookup(columns)
    required = {column["name"] for column in columns if column.get("required", False)}
    best = None
    rows = table.get("rows", [])
    for row_index, row in enumerate(rows[:profile_table["header_search_rows"]]):
        mapping = {}
        duplicates = {}
        for column_index, raw_header in enumerate(row):
            canonical = lookup.get(normalize_header(raw_header))
            if canonical is None:
                continue
            if canonical in mapping:
                duplicates.setdefault(canonical, []).append(column_index)
                continue
            mapping[canonical] = column_index
        score = len(mapping) / len(columns)
        required_score = len(required.intersection(mapping)) / len(required) if required else 1.0
        rank = (required_score, score, -row_index)
        if best is None or rank > best["rank"]:
            best = {
                "rank": rank,
                "row": row_index,
                "mapping": mapping,
                "duplicates": duplicates,
                "score": score,
                "required_score": required_score,
            }
    return best or {
        "rank": (0.0, 0.0, 0),
        "row": 0,
        "mapping": {},
        "duplicates": {},
        "score": 0.0,
        "required_score": 0.0,
    }


def _convert(raw: str, column: dict) -> tuple[object, list[dict]]:
    required = column.get("required", False)
    allow_blank = column.get("allow_blank", not required)
    if raw == "":
        if not allow_blank:
            return None, [_issue(
                "error",
                "blank_required_value",
                f"{column['name']} may not be blank",
                field=column["name"],
            )]
        return None, []

    data_type = column["type"]
    try:
        if data_type == "string":
            value = raw
        elif data_type == "integer":
            if not re.fullmatch(r"[+-]?\d+", raw):
                raise ValueError
            value = int(raw)
        elif data_type == "decimal":
            decimal_value = Decimal(raw)
            if not decimal_value.is_finite():
                raise ValueError
            value = format(decimal_value, "f")
        elif data_type == "boolean":
            normalized = raw.casefold()
            if normalized in {"true", "yes", "y", "1"}:
                value = True
            elif normalized in {"false", "no", "n", "0"}:
                value = False
            else:
                raise ValueError
        else:
            formats = column.get("formats", ["%Y-%m-%d"])
            parsed = next(
                (datetime.strptime(raw, fmt).date() for fmt in formats
                 if _date_matches(raw, fmt)),
                None,
            )
            if parsed is None:
                raise ValueError
            value = parsed.isoformat()
    except (InvalidOperation, ValueError):
        return raw, [_issue(
            "error",
            "invalid_type",
            f"{column['name']} is not a valid {data_type}",
            field=column["name"],
        )]

    issues = []
    comparable = str(value)
    if "pattern" in column and re.fullmatch(column["pattern"], comparable) is None:
        issues.append(_issue(
            "error", "pattern_mismatch", f"{column['name']} does not match its required pattern",
            field=column["name"],
        ))
    if "enum" in column and value not in column["enum"]:
        issues.append(_issue(
            "error", "value_not_allowed", f"{column['name']} is not an allowed value",
            field=column["name"],
        ))
    if data_type in {"integer", "decimal"}:
        numeric = Decimal(str(value))
        if "minimum" in column and numeric < Decimal(str(column["minimum"])):
            issues.append(_issue(
                "error", "below_minimum", f"{column['name']} is below its minimum",
                field=column["name"],
            ))
        if "maximum" in column and numeric > Decimal(str(column["maximum"])):
            issues.append(_issue(
                "error", "above_maximum", f"{column['name']} is above its maximum",
                field=column["name"],
            ))
    return value, issues


def _date_matches(raw: str, fmt: str) -> bool:
    try:
        datetime.strptime(raw, fmt)
        return True
    except ValueError:
        return False


def _is_repeated_header(
    row: list[str], header: list[str], mapping: dict[str, int], lookup: dict[str, str]
) -> bool:
    mapped_indexes = set(mapping.values())
    mapped_match = bool(mapping) and all(
        column_index < len(row)
        and lookup.get(normalize_header(row[column_index])) == canonical
        for canonical, column_index in mapping.items()
    )
    extras_match = all(
        not str(value).strip()
        or (column_index < len(header)
            and normalize_header(value) == normalize_header(header[column_index]))
        for column_index, value in enumerate(row)
        if column_index not in mapped_indexes
    )
    return mapped_match and extras_match


def _evidence(table: dict, row_index: int, column_index: int) -> dict | None:
    rows = table.get("cell_provenance", [])
    if row_index >= len(rows) or column_index >= len(rows[row_index]):
        return None
    return rows[row_index][column_index]


def _extract_records(
    table: dict,
    match: dict,
    profile_table: dict,
    start_index: int,
    headerless_continuation: bool = False,
) -> tuple[list[dict], dict]:
    columns = profile_table["columns"]
    columns_by_name = {column["name"]: column for column in columns}
    lookup = _alias_lookup(columns)
    header_row = match["row"]
    header = table["rows"][header_row] if header_row is not None else []
    mapping = match["mapping"]
    table_issues = []
    source = _source(table)

    if headerless_continuation:
        table_issues.append(_issue(
            "warning",
            "headerless_continuation",
            "adjacent same-width table was extracted using the prior page's column mapping",
            source=source,
        ))

    for warning in table.get("warnings", []):
        table_issues.append(_issue(
            "warning", "extractor_warning", warning, source=source,
        ))
    if table.get("has_merged_cells"):
        table_issues.append(_issue(
            "warning", "merged_cells", "source table contains merged cells", source=source,
        ))
    for canonical, duplicate_indexes in match["duplicates"].items():
        table_issues.append(_issue(
            "error",
            "duplicate_mapped_column",
            f"multiple source columns map to {canonical}",
            field=canonical,
            source={**source, "columns": [mapping[canonical], *duplicate_indexes]},
        ))

    mapped_indexes = set(mapping.values())
    unmapped_headers = [
        {"column": index, "header": value}
        for index, value in enumerate(header)
        if index not in mapped_indexes and str(value).strip()
    ]
    if unmapped_headers and not profile_table["allow_extra_columns"]:
        table_issues.append(_issue(
            "error",
            "extra_columns",
            "source table contains columns not allowed by the profile",
            source=source,
        ))

    records = []
    data_start = 0 if header_row is None else header_row + 1
    for row_index, row in enumerate(table["rows"][data_start:], data_start):
        if not any(str(value).strip() for value in row):
            continue
        if _is_repeated_header(row, header, mapping, lookup):
            continue
        values = {}
        raw_values = {}
        evidence = {}
        record_issues = []
        for canonical, column_index in mapping.items():
            raw = str(row[column_index]).strip() if column_index < len(row) else ""
            value, issues = _convert(raw, columns_by_name[canonical])
            values[canonical] = value
            raw_values[canonical] = raw
            evidence[canonical] = _evidence(table, row_index, column_index)
            for issue in issues:
                record_issues.append({
                    **issue,
                    "source": {
                        **source,
                        "row": row_index,
                        "column": column_index,
                    },
                })
        for column in columns:
            if column["name"] not in mapping:
                values[column["name"]] = None
                raw_values[column["name"]] = ""
                evidence[column["name"]] = None

        unmapped_values = [
            {
                "header": header[column_index] if column_index < len(header) else "",
                "raw": str(row[column_index]).strip() if column_index < len(row) else "",
                "evidence": _evidence(table, row_index, column_index),
            }
            for column_index in range(len(row))
            if column_index not in mapped_indexes
            and (column_index < len(header) and str(header[column_index]).strip()
                 or str(row[column_index]).strip())
        ]
        records.append({
            "index": start_index + len(records),
            "status": "needs_review" if record_issues or table_issues else "accepted",
            "values": values,
            "raw_values": raw_values,
            "evidence": evidence,
            "unmapped_values": unmapped_values,
            "issues": record_issues,
        })

    if not records:
        table_issues.append(_issue(
            "warning",
            "empty_matched_table",
            "table matched the profile but contained no data records",
            source=source,
        ))

    report = {
        "source": source,
        "status": "needs_review" if table_issues else "accepted",
        "header_row": match["row"],
        "header_match": round(match["score"], 3),
        "column_mapping": {
            canonical: {
                "source_column": column_index,
                "source_header": header[column_index] if column_index < len(header) else None,
            }
            for canonical, column_index in mapping.items()
        },
        "unmapped_headers": unmapped_headers,
        "record_count": len(records),
        "issues": table_issues,
    }
    return records, report


def _is_headerless_continuation(previous: tuple[dict, dict] | None, table: dict) -> bool:
    """Conservatively identify a possible PDF continuation and force it through review."""
    if previous is None:
        return False
    previous_table, previous_match = previous
    previous_page = previous_table.get("page")
    current_page = table.get("page")
    previous_width = previous_table.get("column_count")
    current_width = table.get("column_count")
    return (
        previous_page is not None
        and current_page == previous_page + 1
        and table.get("index") == 0
        and previous_width is not None
        and previous_width == current_width
        and all(index < current_width for index in previous_match["mapping"].values())
    )


def extract_with_profile(path: str, profile: str | dict = "lab-coa-v1") -> dict:
    """Extract only rows that satisfy a versioned profile and return an auditable decision."""
    checked_profile = load_profile(profile)
    profile_table = checked_profile["table"]
    resolved, document = _document(path)
    if document["type"] == "pdf":
        extraction = extractor.extract_tables(resolved, merge_multipage=False)
        engine = {"name": "pdfplumber", "version": _package_version("pdfplumber")}
    else:
        extraction = docx_extractor.extract_docx_tables(resolved)
        engine = {"name": "python-docx", "version": _package_version("python-docx")}

    required = {
        column["name"] for column in profile_table["columns"] if column.get("required", False)
    }
    candidate_reports = []
    records = []
    matched_count = 0
    compatible_count = 0
    continuation_count = 0
    previous_compatible = None
    for table in extraction["tables"]:
        match = _header_match(table, profile_table)
        source = _source(table)
        if match["score"] < profile_table["minimum_header_match"]:
            if _is_headerless_continuation(previous_compatible, table):
                inherited_match = {
                    "row": None,
                    "mapping": dict(previous_compatible[1]["mapping"]),
                    "duplicates": {},
                    "score": 0.0,
                }
                table_records, report = _extract_records(
                    table,
                    inherited_match,
                    profile_table,
                    len(records),
                    headerless_continuation=True,
                )
                records.extend(table_records)
                candidate_reports.append(report)
                continuation_count += 1
                previous_compatible = (table, inherited_match)
                continue
            partial_issues = []
            if match["score"] > 0:
                partial_issues.append(_issue(
                    "warning",
                    "partial_profile_match",
                    "table matched some profile headers but not the configured threshold",
                    source=source,
                ))
            candidate_reports.append({
                "source": source,
                "status": "needs_review" if partial_issues else "not_matched",
                "header_row": match["row"],
                "header_match": round(match["score"], 3),
                "record_count": 0,
                "issues": partial_issues,
            })
            if previous_compatible and table.get("page") not in {
                previous_compatible[0].get("page"),
                previous_compatible[0].get("page", 0) + 1,
            }:
                previous_compatible = None
            continue
        matched_count += 1
        missing = sorted(required.difference(match["mapping"]))
        if missing:
            candidate_reports.append({
                "source": source,
                "status": "rejected",
                "header_row": match["row"],
                "header_match": round(match["score"], 3),
                "record_count": 0,
                "issues": [_issue(
                    "error",
                    "missing_required_columns",
                    f"required columns were not found: {', '.join(missing)}",
                    source=source,
                )],
            })
            previous_compatible = None
            continue
        compatible_count += 1
        table_records, report = _extract_records(table, match, profile_table, len(records))
        records.extend(table_records)
        candidate_reports.append(report)
        previous_compatible = (table, match)

    document_issues = []
    if not matched_count:
        document_issues.append(_issue(
            "error",
            "no_matching_table",
            "no table met the profile's header-match threshold",
        ))
    elif not compatible_count:
        document_issues.append(_issue(
            "error",
            "no_compatible_table",
            "matched tables were missing required columns",
        ))
    if len(records) < profile_table["min_records"]:
        document_issues.append(_issue(
            "error",
            "insufficient_records",
            f"profile requires at least {profile_table['min_records']} record(s)",
        ))

    unique_by = profile_table["unique_by"]
    if unique_by:
        seen = {}
        for record in records:
            key = tuple(record["values"].get(field) for field in unique_by)
            if any(value is None for value in key):
                continue
            if key in seen:
                issue = _issue(
                    "error",
                    "duplicate_record_key",
                    f"duplicate record key for {', '.join(unique_by)}",
                    records=[seen[key], record["index"]],
                )
                document_issues.append(issue)
                for duplicate_index in (seen[key], record["index"]):
                    records[duplicate_index]["issues"].append(issue)
                    records[duplicate_index]["status"] = "needs_review"
            else:
                seen[key] = record["index"]

    collected_issues = [
        *document_issues,
        *(issue for table in candidate_reports for issue in table.get("issues", [])),
        *(issue for record in records for issue in record["issues"]),
    ]
    all_issues = []
    seen_issues = set()
    for issue in collected_issues:
        serialized = json.dumps(issue, sort_keys=True, separators=(",", ":"))
        if serialized not in seen_issues:
            seen_issues.add(serialized)
            all_issues.append(issue)
    fatal_codes = {"no_matching_table", "no_compatible_table", "insufficient_records"}
    if any(issue["code"] in fatal_codes for issue in document_issues):
        decision = "rejected"
    elif all_issues:
        decision = "needs_review"
    else:
        decision = "accepted"

    checked_profile_hash = profile_sha256(checked_profile)
    fingerprint_basis = json.dumps({
        "document_sha256": document["sha256"],
        "profile_sha256": checked_profile_hash,
        "pdf_agent_mcp_version": __version__,
        "engine": engine,
        "result_schema_version": RESULT_SCHEMA_VERSION,
    }, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    fingerprint = hashlib.sha256(fingerprint_basis.encode("ascii")).hexdigest()
    return {
        "$schema": "https://raw.githubusercontent.com/wesseltl/pdf-mcp/main/extraction-result.schema.json",
        "result_schema_version": RESULT_SCHEMA_VERSION,
        "decision": decision,
        "review_required": decision != "accepted",
        "profile": {
            "id": checked_profile["id"],
            "version": checked_profile["version"],
            "description": checked_profile["description"],
            "sha256": checked_profile_hash,
            "fields": [column["name"] for column in profile_table["columns"]],
        },
        "document": document,
        "records": records,
        "tables": candidate_reports,
        "summary": {
            "detected_tables": extraction["n_tables"],
            "matched_tables": matched_count,
            "compatible_tables": compatible_count,
            "headerless_continuations": continuation_count,
            "records": len(records),
            "accepted_records": sum(record["status"] == "accepted" for record in records),
            "issues": len(all_issues),
        },
        "review": {"issues": all_issues},
        "audit": {
            "extraction_fingerprint": fingerprint,
            "pdf_agent_mcp_version": __version__,
            "engine": engine,
        },
    }


def _issue_text(issues: list[dict]) -> str:
    return "; ".join(f"{issue['code']}: {issue['message']}" for issue in issues)


def _write_xlsx(result: dict, output_path: str) -> None:
    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    bold = Font(bold=True)
    fills = {
        "accepted": PatternFill("solid", fgColor="E2F0D9"),
        "needs_review": PatternFill("solid", fgColor="FFF2CC"),
        "rejected": PatternFill("solid", fgColor="F4CCCC"),
    }
    metadata = [
        ("decision", result["decision"]),
        ("profile", result["profile"]["id"]),
        ("profile_version", result["profile"]["version"]),
        ("document_sha256", result["document"]["sha256"]),
        ("extraction_fingerprint", result["audit"]["extraction_fingerprint"]),
        ("records", result["summary"]["records"]),
        ("issues", result["summary"]["issues"]),
    ]
    for key, value in metadata:
        review.append([key, value])
    for row in review.iter_rows(min_row=1, max_col=1):
        row[0].font = bold
    for cell in review[1]:
        cell.fill = fills[result["decision"]]
    review.append([])
    review.append(["severity", "code", "message", "source"])
    for cell in review[review.max_row]:
        cell.font = bold
    for issue in result["review"]["issues"]:
        review.append([
            issue["severity"], issue["code"], issue["message"],
            json.dumps(issue.get("source", {}), sort_keys=True),
        ])
    review.freeze_panes = "A10"

    data = workbook.create_sheet("Data")
    fields = result["profile"]["fields"]
    data.append(["record", "status", *[spreadsheet_safe(field) for field in fields], "issues"])
    for cell in data[1]:
        cell.font = bold
    for record in result["records"]:
        data.append([
            record["index"] + 1,
            record["status"],
            *[spreadsheet_safe(record["values"].get(field)) for field in fields],
            _issue_text(record["issues"]),
        ])
        for cell in data[data.max_row]:
            cell.fill = fills[record["status"]]
    data.freeze_panes = "A2"

    evidence = workbook.create_sheet("Evidence")
    evidence.append([
        "record", "field", "raw", "value", "page", "table", "row", "column", "bbox"
    ])
    for cell in evidence[1]:
        cell.font = bold
    for record in result["records"]:
        for field in fields:
            source = record["evidence"].get(field) or {}
            evidence.append([
                record["index"] + 1,
                spreadsheet_safe(field),
                spreadsheet_safe(record["raw_values"].get(field)),
                spreadsheet_safe(record["values"].get(field)),
                source.get("page"),
                source.get("table_index"),
                source.get("row"),
                source.get("column"),
                json.dumps(source.get("bbox")) if source.get("bbox") is not None else "",
            ])
    evidence.freeze_panes = "A2"
    workbook.save(output_path)


def _write_csv(result: dict, output_path: str) -> None:
    fields = result["profile"]["fields"]
    with open(output_path, "w", newline="", encoding="utf-8") as output:
        writer = csv.writer(output)
        writer.writerow(["decision", result["decision"]])
        writer.writerow(["profile", result["profile"]["id"]])
        writer.writerow(["document_sha256", result["document"]["sha256"]])
        writer.writerow([])
        writer.writerow([
            "record", "status", *[spreadsheet_safe(field) for field in fields], "issues"
        ])
        for record in result["records"]:
            writer.writerow([
                record["index"] + 1,
                record["status"],
                *[spreadsheet_safe(record["values"].get(field)) for field in fields],
                _issue_text(record["issues"]),
            ])


def export_with_profile(input_path: str, profile: str | dict, output_path: str) -> dict:
    """Write a profile-checked extraction to evidence-rich XLSX, CSV, or JSON."""
    result = extract_with_profile(input_path, profile)
    resolved_output = extractor._check_path(output_path)
    output_type = Path(resolved_output).suffix.lower()
    if output_type not in {".xlsx", ".csv", ".json"}:
        raise ValueError("profile output must be .xlsx, .csv, or .json")
    os.makedirs(os.path.dirname(resolved_output) or ".", exist_ok=True)
    if output_type == ".xlsx":
        _write_xlsx(result, resolved_output)
    elif output_type == ".csv":
        _write_csv(result, resolved_output)
    else:
        with open(resolved_output, "w", encoding="utf-8") as output:
            json.dump(result, output, indent=2)
            output.write("\n")
    return {
        "output": resolved_output,
        "output_type": output_type[1:],
        "decision": result["decision"],
        "records": result["summary"]["records"],
        "issues": result["summary"]["issues"],
        "profile": result["profile"],
        "document_sha256": result["document"]["sha256"],
        "extraction_fingerprint": result["audit"]["extraction_fingerprint"],
    }
