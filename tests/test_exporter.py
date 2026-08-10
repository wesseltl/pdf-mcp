"""Tests for exporting extracted tables to user-facing files."""
import json
import os
import sys
import tempfile
import unittest
from io import StringIO
from unittest import mock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pdf_mcp import cli, exporter


def make_pdf(path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    doc = SimpleDocTemplate(path, pagesize=A4)
    table = Table([["Item", "Qty"], ["Widget", "3"], ["Bolt", "10"]])
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    doc.build([table])


def make_docx(path):
    from docx import Document
    doc = Document()
    table = doc.add_table(rows=1, cols=2)
    table.rows[0].cells[0].text = "Analyte"
    table.rows[0].cells[1].text = "Result"
    for analyte, result in [("pH", "7.2"), ("Sodium", "140")]:
        cells = table.add_row().cells
        cells[0].text = analyte
        cells[1].text = result
    doc.save(path)


class TestExporter(unittest.TestCase):
    def test_export_pdf_to_xlsx(self):
        from openpyxl import load_workbook
        tmp = tempfile.mkdtemp()
        input_path = os.path.join(tmp, "invoice.pdf")
        output_path = os.path.join(tmp, "invoice.xlsx")
        make_pdf(input_path)

        result = exporter.export_document_tables(input_path, output_path)

        self.assertEqual(result["output_type"], "xlsx")
        self.assertEqual(result["n_tables"], 1)
        self.assertTrue(os.path.exists(output_path))
        wb = load_workbook(output_path)
        self.assertEqual(wb.sheetnames, ["Review", "Table 1"])
        self.assertEqual(wb["Review"]["A1"].value, "table")
        self.assertEqual(wb["Table 1"]["A6"].value, "Item")
        self.assertEqual(wb["Table 1"]["A7"].value, "Widget")

    def test_export_docx_to_csv(self):
        tmp = tempfile.mkdtemp()
        input_path = os.path.join(tmp, "coa.docx")
        output_path = os.path.join(tmp, "coa.csv")
        make_docx(input_path)

        result = exporter.export_document_tables(input_path, output_path)

        self.assertEqual(result["output_type"], "csv")
        with open(output_path, encoding="utf-8") as f:
            content = f.read()
        self.assertIn("source_type,docx", content)
        self.assertIn("Analyte,Result", content)
        self.assertIn("Sodium,140", content)

    def test_export_docx_to_json(self):
        tmp = tempfile.mkdtemp()
        input_path = os.path.join(tmp, "coa.docx")
        output_path = os.path.join(tmp, "coa.json")
        make_docx(input_path)

        exporter.export_document_tables(input_path, output_path)

        with open(output_path, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(data["source_type"], "docx")
        self.assertEqual(data["tables"][0]["rows"][0], ["Analyte", "Result"])

    def test_export_rejects_unsupported_input(self):
        with self.assertRaises(ValueError):
            exporter.export_document_tables("sample.txt", "sample.xlsx")

    def test_cli_exports_document_tables(self):
        tmp = tempfile.mkdtemp()
        input_path = os.path.join(tmp, "coa.docx")
        output_path = os.path.join(tmp, "coa.json")
        make_docx(input_path)

        with mock.patch.object(sys, "argv", ["export-document-tables", input_path, output_path]), \
                mock.patch("sys.stdout", new_callable=StringIO):
            self.assertEqual(cli.main(), 0)
        self.assertTrue(os.path.exists(output_path))


if __name__ == "__main__":
    unittest.main()
