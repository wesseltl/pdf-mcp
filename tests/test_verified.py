"""Tests for profile-checked extraction, evidence, exports, and evaluation."""
import json
import os
import tempfile
import unittest
from io import StringIO
from unittest import mock

from pdf_mcp import evaluator, profile_cli, verified


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def make_pdf(path, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    table = Table(rows)
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    SimpleDocTemplate(path, pagesize=A4).build([table])


def make_two_table_pdf(path, first_rows, second_rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

    def table(rows):
        result = Table(rows)
        result.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
        return result

    SimpleDocTemplate(path, pagesize=A4).build([
        table(first_rows), Spacer(1, 36), table(second_rows),
    ])


def make_merged_docx(path):
    from docx import Document
    document = Document()
    table = document.add_table(rows=3, cols=3)
    table.rows[0].cells[0].merge(table.rows[0].cells[2]).text = "Results"
    for column, value in enumerate(["Analyte", "Result", "Unit"]):
        table.rows[1].cells[column].text = value
    for column, value in enumerate(["pH", "7.2", ""]):
        table.rows[2].cells[column].text = value
    document.save(path)


def make_multipage_lab_pdf(path, repeat_header):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import PageBreak, SimpleDocTemplate, Table, TableStyle

    def table(rows):
        result = Table(rows)
        result.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
        return result

    second = [["Analyte", "Result"], ["Sodium", "140"], ["Potassium", "4.2"]]
    if not repeat_header:
        second = second[1:]
    SimpleDocTemplate(path, pagesize=A4).build([
        table([["Analyte", "Result"], ["pH", "7.2"], ["Glucose", "5.1"]]),
        PageBreak(),
        table(second),
    ])


class VerifiedExtractionTests(unittest.TestCase):
    def test_invoice_profile_is_accepted_with_cell_evidence(self):
        result = verified.extract_with_profile(
            os.path.join(ROOT, "examples", "invoice.pdf"), "invoice-lines-v1"
        )
        self.assertEqual(result["decision"], "accepted")
        self.assertEqual(result["result_schema_version"], "1.0")
        self.assertEqual(result["summary"]["records"], 3)
        self.assertEqual(result["records"][0]["values"]["unit_price"], "12.50")
        evidence = result["records"][0]["evidence"]["item"]
        self.assertEqual(evidence["page"], 1)
        self.assertEqual(len(evidence["bbox"]), 4)
        self.assertEqual(len(result["audit"]["extraction_fingerprint"]), 64)

    def test_local_mcp_exposes_profile_contract(self):
        from pdf_mcp import server
        self.assertIn("lab-coa-v1", {
            profile["id"] for profile in server.list_extraction_profiles()
        })
        result = server.extract_with_profile(
            os.path.join(ROOT, "examples", "invoice.pdf"), "invoice-lines-v1"
        )
        self.assertEqual(result["decision"], "accepted")

    def test_wrong_profile_fails_closed(self):
        result = verified.extract_with_profile(
            os.path.join(ROOT, "examples", "invoice.pdf"), "lab-coa-v1"
        )
        self.assertEqual(result["decision"], "rejected")
        self.assertEqual(result["records"], [])
        self.assertTrue(result["review_required"])
        self.assertIn("no_matching_table", {issue["code"] for issue in result["review"]["issues"]})

    def test_invalid_typed_value_is_routed_to_review_and_preserved(self):
        path = os.path.join(tempfile.mkdtemp(), "invalid.pdf")
        make_pdf(path, [["Item", "Qty", "Price"], ["Widget", "many", "12.50"]])
        result = verified.extract_with_profile(path, "invoice-lines-v1")
        self.assertEqual(result["decision"], "needs_review")
        record = result["records"][0]
        self.assertEqual(record["raw_values"]["quantity"], "many")
        self.assertEqual(record["values"]["quantity"], "many")
        self.assertIn("invalid_type", {issue["code"] for issue in record["issues"]})

    def test_non_finite_decimal_is_routed_to_review(self):
        path = os.path.join(tempfile.mkdtemp(), "invalid.pdf")
        make_pdf(path, [["Item", "Qty", "Price"], ["Widget", "NaN", "12.50"]])
        result = verified.extract_with_profile(path, "invoice-lines-v1")
        self.assertEqual(result["decision"], "needs_review")
        self.assertIn(
            "invalid_type",
            {issue["code"] for issue in result["records"][0]["issues"]},
        )

    def test_decimal_values_have_a_stable_non_exponent_form(self):
        path = os.path.join(tempfile.mkdtemp(), "decimal.pdf")
        make_pdf(path, [["Item", "Qty", "Price"], ["Widget", "1e3", "12.50"]])
        result = verified.extract_with_profile(path, "invoice-lines-v1")
        self.assertEqual(result["decision"], "accepted")
        self.assertEqual(result["records"][0]["values"]["quantity"], "1000")

    def test_merged_docx_is_never_silently_accepted(self):
        path = os.path.join(tempfile.mkdtemp(), "merged.docx")
        make_merged_docx(path)
        result = verified.extract_with_profile(path, "lab-coa-v1")
        self.assertEqual(result["decision"], "needs_review")
        self.assertTrue(any(issue["code"] == "extractor_warning"
                            for issue in result["review"]["issues"]))

    def test_headerless_multipage_continuation_is_extracted_and_forces_review(self):
        path = os.path.join(tempfile.mkdtemp(), "multipage.pdf")
        make_multipage_lab_pdf(path, repeat_header=False)
        result = verified.extract_with_profile(path, "lab-coa-v1")
        self.assertEqual(result["decision"], "needs_review")
        self.assertEqual(result["summary"]["records"], 4)
        self.assertEqual(result["summary"]["headerless_continuations"], 1)
        self.assertEqual(result["records"][-1]["values"]["analyte"], "Potassium")
        self.assertIn(
            "headerless_continuation",
            {issue["code"] for issue in result["review"]["issues"]},
        )

    def test_repeated_multipage_header_can_be_accepted_without_stitching(self):
        path = os.path.join(tempfile.mkdtemp(), "multipage.pdf")
        make_multipage_lab_pdf(path, repeat_header=True)
        result = verified.extract_with_profile(path, "lab-coa-v1")
        self.assertEqual(result["decision"], "accepted")
        self.assertEqual(result["summary"]["records"], 4)

    def test_partial_profile_match_prevents_false_acceptance(self):
        path = os.path.join(tempfile.mkdtemp(), "partial.pdf")
        make_two_table_pdf(
            path,
            [["Item", "Qty", "Price"], ["Widget", "1", "5.00"]],
            [["Description", "Code"], ["Shipping", "S"]],
        )
        result = verified.extract_with_profile(path, "invoice-lines-v1")
        self.assertEqual(result["decision"], "needs_review")
        self.assertIn(
            "partial_profile_match",
            {issue["code"] for issue in result["review"]["issues"]},
        )

    def test_xlsx_contains_review_data_and_evidence_and_escapes_formula(self):
        temp = tempfile.mkdtemp()
        input_path = os.path.join(temp, "formula.pdf")
        output_path = os.path.join(temp, "formula.xlsx")
        make_pdf(input_path, [["Item", "Qty", "Price"], ["=2+3", "1", "5.00"]])
        result = verified.export_with_profile(
            input_path, "invoice-lines-v1", output_path
        )
        self.assertEqual(result["decision"], "accepted")

        from openpyxl import load_workbook
        workbook = load_workbook(output_path, data_only=False)
        self.assertEqual(workbook.sheetnames, ["Review", "Data", "Evidence"])
        self.assertEqual(workbook["Review"]["B1"].value, "accepted")
        self.assertEqual(workbook["Data"]["C2"].value, "'=2+3")
        self.assertEqual(workbook["Evidence"]["C2"].value, "'=2+3")

    def test_extract_cli_uses_nonzero_exit_for_review(self):
        temp = tempfile.mkdtemp()
        input_path = os.path.join(temp, "invalid.pdf")
        output_path = os.path.join(temp, "result.json")
        make_pdf(input_path, [["Item", "Qty", "Price"], ["Widget", "many", "12.50"]])
        with mock.patch("sys.stdout", new_callable=StringIO):
            status = profile_cli.extract_main(
                [input_path, "invoice-lines-v1", output_path]
            )
        self.assertEqual(status, 2)
        with open(output_path, encoding="utf-8") as result_file:
            self.assertEqual(json.load(result_file)["decision"], "needs_review")

    def test_sample_evaluation_passes_without_exposing_cells(self):
        report = evaluator.evaluate_manifest(
            os.path.join(ROOT, "evaluations", "sample-invoice.json")
        )
        self.assertTrue(report["passed"])
        self.assertEqual(report["evaluation_report_schema_version"], "1.0")
        self.assertEqual(report["metrics"]["field_f1"], 1.0)
        self.assertEqual(len(report["cases"][0]["extraction_fingerprint"]), 64)
        self.assertNotIn("Widget", json.dumps(report))

    def test_evaluation_threshold_failure_returns_nonzero_without_exposing_cells(self):
        temp = tempfile.mkdtemp()
        manifest_path = os.path.join(temp, "evaluation.json")
        with open(manifest_path, "w", encoding="utf-8") as manifest:
            json.dump({
                "evaluation_schema_version": "1.0",
                "evidence_label": "synthetic_regression",
                "profile": "invoice-lines-v1",
                "minimums": {"field_f1": 1.0},
                "cases": [{
                    "id": "wrong-ground-truth",
                    "document": os.path.join(ROOT, "examples", "invoice.pdf"),
                    "expected_decision": "accepted",
                    "expected_records": [
                        {"item": "Incorrect", "quantity": "3", "unit_price": "12.50"}
                    ],
                }],
            }, manifest)
        report = evaluator.evaluate_manifest(manifest_path)
        self.assertFalse(report["passed"])
        self.assertEqual(report["evidence_label"], "synthetic_regression")
        self.assertNotIn("Incorrect", json.dumps(report))
        with mock.patch("sys.stdout", new_callable=StringIO):
            self.assertEqual(profile_cli.evaluate_main([manifest_path]), 1)


if __name__ == "__main__":
    unittest.main()
